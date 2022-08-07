import encodings
from itertools import chain
import string
import random
from io import BytesIO as IO
from urllib import response
from django.views.decorators.cache import never_cache
from django.core.mail import send_mail, EmailMessage
from django.http import HttpResponseRedirect, HttpResponse
from django.db import connection
import csv
import xlwt
from django.shortcuts import render
from django.template.loader import get_template
from pandas.io.sas.sas_constants import magic
import openpyxl
from xhtml2pdf import pisa
import pandas as pd
import datetime
from student.models import User
# Create your views here.
from student.utils import render_to_pdf
from tnp_admin.models import Admin, StudentsEligible, StudentPlaced,mailResponse
from tnp_admin.models import Company
from secrets import token_urlsafe
from student.models import Resume
from django.db.models import Q
import urllib.parse

@never_cache
def dashboard(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        if (request.session['admin_type'] == 'Super'):
            if request.method == "POST":
                psw = request.POST['psw']
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']
                # ,admin_type=admin_type
                Admin.objects.filter(username=admin,admin_type=admin_type).update(password=psw)
                data = Admin.objects.filter(username=admin,admin_type=admin_type)

                return render(request, 'adminDashb.html', {'data': data, 'msg': "Password Updated.",admin_type:'super'})
            else:
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']

                # admin_type=request.session['admin_type'],admin_type=admin_type
                data = Admin.objects.filter(username=admin,admin_type=admin_type)
                # dir(data)
                # print(data)
                return render(request, 'adminDashb.html', {'data': data})
        elif (request.session['admin_type'] == 'Information Technology'):
            if request.method == "POST":
                psw = request.POST['psw']
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']
                # ,admin_type=admin_type

                Admin.objects.filter(username=admin,admin_type=admin_type).update(password=psw)
                data = Admin.objects.filter(username=admin,admin_type=admin_type)

                return render(request, 'adminDashb.html', {'data': data, 'msg': "Password Updated."})
            else:
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']

                # admin_type=request.session['admin_type'],admin_type=admin_type
                data = Admin.objects.filter(username=admin,admin_type=admin_type)
                # dir(data)
                # print(data)
                return render(request, 'adminDashb.html', {'data': data})
        elif (request.session['admin_type'] == 'Computer'):
            if request.method == "POST":
                psw = request.POST['psw']
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']
                # ,admin_type=admin_type

                Admin.objects.filter(username=admin,admin_type=admin_type).update(password=psw)
                data = Admin.objects.filter(username=admin,admin_type=admin_type)

                return render(request, 'adminDashb.html', {'data': data, 'msg': "Password Updated."})
            else:
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']

                # admin_type=request.session['admin_type'],admin_type=admin_type
                data = Admin.objects.filter(username=admin,admin_type=admin_type)
                # dir(data)
                # print(data)
                return render(request, 'adminDashb.html', {'data': data})
        elif (request.session['admin_type'] == 'Electronics & Telecommunication'):
            if request.method == "POST":
                psw = request.POST['psw']
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']
                # ,admin_type=admin_type

                Admin.objects.filter(username=admin,admin_type=admin_type).update(password=psw)
                data = Admin.objects.filter(username=admin,admin_type=admin_type)

                return render(request, 'adminDashb.html', {'data': data, 'msg': "Password Updated."})
            else:
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']

                # admin_type=request.session['admin_type'],admin_type=admin_type
                data = Admin.objects.filter(username=admin,admin_type=admin_type)
                # dir(data)
                # print(data)
                return render(request, 'adminDashb.html', {'data': data})
        elif (request.session['admin_type'] == 'Electronics'):
            if request.method == "POST":
                psw = request.POST['psw']
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']
                # ,admin_type=admin_type

                Admin.objects.filter(username=admin,admin_type=admin_type).update(password=psw)
                data = Admin.objects.filter(username=admin,admin_type=admin_type)

                return render(request, 'adminDashb.html', {'data': data, 'msg': "Password Updated."})
            else:
                admin = request.session['admin_username']
                admin_type=request.session['admin_type']

                # admin_type=request.session['admin_type'],admin_type=admin_type
                data = Admin.objects.filter(username=admin,admin_type=admin_type)
                # dir(data)
                # print(data)
                return render(request, 'adminDashb.html', {'data': data})


@never_cache
def display(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_types=request.session.get('admin_type')
        if (admin_types=="Super"):
            resumes = Resume.objects.all()
            users = User.objects.all()
            temp = []
            if resumes.count() > 0 or users.count() > 0:
                for user in users:
                    flag = 0
                    for resume in resumes:
                        if user.username == resume.somaiya_id:
                            flag = 0
                            break
                        else:
                            flag = 1
                    if flag == 1:
                        temp.append(user.username)
                    else:
                        pass
                return render(request, 'display_student.html', {'userDetails': resumes, 'temp': temp, 'users': users})
            elif resumes.count() == 0:
                return render(request, 'display_student.html', {'users': users, 'userDetails': resumes})
            else:
                return render(request, 'display_student.html')
        if (admin_types=="Information Technology"):
            resumes = Resume.objects.filter(branch=admin_types)
            users = User.objects.filter(branch=admin_types)
            temp = []
            if resumes.count() > 0 or users.count() > 0:
                for user in users:
                    flag = 0
                    for resume in resumes:
                        if user.username == resume.somaiya_id:
                            flag = 0
                            break
                        else:
                            flag = 1
                    if flag == 1:
                        temp.append(user.username)
                    else:
                        pass
                return render(request, 'display_student.html', {'userDetails': resumes, 'temp': temp, 'users': users})
            elif resumes.count() == 0:
                return render(request, 'display_student.html', {'users': users, 'userDetails': resumes})
            else:
                return render(request, 'display_student.html')
        if (admin_types=="Computer"):
            resumes = Resume.objects.filter(branch=admin_types)
            users = User.objects.filter(branch=admin_types)
            temp = []
            if resumes.count() > 0 or users.count() > 0:
                for user in users:
                    flag = 0
                    for resume in resumes:
                        if user.username == resume.somaiya_id:
                            flag = 0
                            break
                        else:
                            flag = 1
                    if flag == 1:
                        temp.append(user.username)
                    else:
                        pass
                return render(request, 'display_student.html', {'userDetails': resumes, 'temp': temp, 'users': users})
            elif resumes.count() == 0:
                return render(request, 'display_student.html', {'users': users, 'userDetails': resumes})
            else:
                return render(request, 'display_student.html')
        if (admin_types=="Electronics & Telecommunication"):
            resumes = Resume.objects.filter(branch=admin_types)
            users = User.objects.filter(branch=admin_types)
            temp = []
            if resumes.count() > 0 or users.count() > 0:
                for user in users:
                    flag = 0
                    for resume in resumes:
                        if user.username == resume.somaiya_id:
                            flag = 0
                            break
                        else:
                            flag = 1
                    if flag == 1:
                        temp.append(user.username)
                    else:
                        pass
                return render(request, 'display_student.html', {'userDetails': resumes, 'temp': temp, 'users': users})
            elif resumes.count() == 0:
                return render(request, 'display_student.html', {'users': users, 'userDetails': resumes})
            else:
                return render(request, 'display_student.html')
        if (admin_types=="Electronics"):
            resumes = Resume.objects.filter(branch=admin_types)
            users = User.objects.filter(branch=admin_types)
            temp = []
            if resumes.count() > 0 or users.count() > 0:
                for user in users:
                    flag = 0
                    for resume in resumes:
                        if user.username == resume.somaiya_id:
                            flag = 0
                            break
                        else:
                            flag = 1
                    if flag == 1:
                        temp.append(user.username)
                    else:
                        pass
                return render(request, 'display_student.html', {'userDetails': resumes, 'temp': temp, 'users': users})
            elif resumes.count() == 0:
                return render(request, 'display_student.html', {'users': users, 'userDetails': resumes})
            else:
                return render(request, 'display_student.html')


@never_cache
def add_admin(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        if request.method == "POST":
            name = request.POST['name']
            username = request.POST['uname'].strip()
            password = str(''.join(random.choices(string.ascii_uppercase +
                                string.digits, k = 8)))
            branch = request.POST['branch']
            admin_type=request.POST['branch']

            check = Admin.objects.filter(username=username)
            if check:
                msg = {
                        'invalidate': "User already exists."
                }
                return render(request, 'add_admin.html', msg)
            else:
                addUser = Admin(username=username, password=password, dept=branch, role="TNP Admin", name=name,admin_type=admin_type)
                addUser.save()
                # send_mail(
                #     'Welcome to Placement Information Portal \n Login //tnpportal.kjsieit.in/',
                #     'Id: ' + username + '\nPassword: ' + password + '.',
                #     'tnpportal7@gmail.com',
                #      [username],
                #      fail_silently=False,
                #  )
                email_boby = """
                <html>
                <body>
                    <h2>Welcome To Placement Information Portal Admin of department %s</h2>
                <p>Follow up is the credentials for login in to PIMS Portal </p>
                <p><strong>Username:</strong> %s</p>
                <p><strong>password:</strong> %s</p>
                <p>Link of the website is <a href="//tnpportal.kjsieit.in/">Pims Portal</a>

                </body>
                </html>
                """%(admin_type,username,password)
                
                email = EmailMessage('Placement',email_boby,"tnpportal7@gmail.com",[username])
                email.content_subtype = 'html'
                email.send()
                msg = {
                    'validate': "Added successfully."
                }
                return render(request, 'add_admin.html', msg)
        else:
            msg={
                'validate': "Add Admin"
            }
            return render(request, 'add_admin.html',msg)
        # else:
        #     msg={
        #         'validate' : "You cannot add admin."
        #     }
        #     return render(request, 'add_admin.html', msg)



@never_cache
def add_user(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        if request.method == "POST":
            name = request.POST['name']
            username = request.POST['uname'].strip()
            password = str(''.join(random.choices(string.ascii_uppercase +
                             string.digits, k = 8)))
            # password = request.POST['psw']
            branch = request.POST['branch']

            check = User.objects.filter(username=username)
            if check:
                msg = {
                    'invalidate': "User already exists."
                }
                return render(request, 'add_student.html', msg)
            else:
                addUser = User(name=name, username=username, password=password, branch=branch)
                addUser.save()
                # send_mail(
                #      'Placement Portal',
                #      'Id: ' + username + '\nPassword: ' + password + '.',
                #      'tnpportal7@gmail.com',
                #      [username],
                #      fail_silently=False,
                #  )
                email_boby = """
                <html>
                <body>
                    <h2>Welcome To Placement Information Portal Students</h2>
                <p>Follow up is the credentials for login in to PIMS Portal </p><br>
                <p>Link of the website is <a href="//tnpportal.kjsieit.in/">Pims Portal</a>
                <p><strong>Username:</strong> %s</p>
                <p><strong>password:</strong> %s</p><br>
                </body>
                </html>
                """%(username,password)
                
                email = EmailMessage('Placement',email_boby,"tnpportal7@gmail.com",[username])
                email.content_subtype = 'html'
                email.send()
                msg = {
                    'validate': "Added successfully."
                }
                return render(request, 'add_student.html', msg)
        else:
            return render(request, 'add_student.html')


@never_cache
def add_company(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        if request.method == "POST":
            types=request.session.get('admin_type')
            c_name = request.POST['c_name'].strip()
            compCheck = Company.objects.filter(comp_name=c_name)

            if not compCheck.exists():
                c_profile = request.POST['c_profile']
                ctc = request.POST['ctc']
                eligible = request.POST['eligible']
                bond = request.POST['bond']
                date = request.POST['date']
                time = request.POST['time']
                venue = request.POST['venue']
                any_live_kt = request.POST['live_kt']
                print('live',any_live_kt)
                branch = request.POST.getlist('branch')
                instruction = request.POST['instruction']
                campus = request.POST['campus']
                studentObj = Resume.objects.filter(agg__gte=eligible, branch__in=branch)
                # print("eleibigle",type(eligible),studentObj)

                branch = ','.join(map(str, branch))

                # addCompany = Company(comp_name=c_name, comp_profile=c_profile, ctc=ctc, eligibility=eligible, bond=bond,
                #                          date=date, time=time, venue=venue, branch=branch, instruction=instruction, campus=campus)
                # addCompany.save()

                temp = []
                # and student.oneto6 == ""
                # if studentObj.count() > 0 :
                #     for student in studentObj:
                #         if int(ctc) <= 600000 and student.oneto6 == "":
                #             temp.append(student.somaiya_id)
                #         elif int(ctc) > 600000:
                #             temp.append(student.somaiya_id)
                #         else:
                #             pass
                    # return render(request, 'check_eligible.html',
                    #               {'temp': temp, 'c_name': c_name, 'studentObj': studentObj})
                # else:
                #     msg = {
                #         'success': "No student eligible."
                #     }
                return render(request, 'add_company.html')
            else:
                msg = {
                    'success': "Company already exists."
                }
                return render(request, 'add_company.html', msg)
        else:
            return render(request, 'add_company.html')


@never_cache
def check_eligible(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        types=request.session.get('admin_type')
        if (types=="Super"):
            company = request.POST['company']
            student = request.POST['hidden']
            comp = Company.objects.get(comp_name=company)

            # print(student)
            stud = student.split(",")
            stud_arr = []

            for temp in stud:
                if temp != "on":
                    user = temp
                    studentObj = Resume.objects.get(somaiya_id=user)
                    name = studentObj.name
                    branch = studentObj.branch
                    studentEligible = StudentsEligible(stud_user=user, comp_name=comp.comp_name, stud_name=name,
                                                        branch=branch)
                    studentEligible.save()
                    stud_arr.append(temp)
                    current_time = datetime.datetime.now()
                    token = token_urlsafe(16)
                    saveToken = mailResponse(comp_name=comp.comp_name,stud_user=user,token=token,time=current_time, stud_response=0)
                    saveToken.save()
                    time = str(comp.time)
                    date = str(comp.date)
                    email_body = """\
            <html>
            <head></head>
            <body>
                <p>Congratulations. You are eligible for %s.</p>
                <p>Please find below the details for the company.</p>
                <table style="border: 1px solid #dddddd;">
                    <tr>
                        <th style="border: 1px solid #6A6969;">Company Name</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Profile</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">CTC</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Branch</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Eligibility</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Date</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Time</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Venue</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Bond</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">Instruction</th>
                        <td style="border: 1px solid #6A6969;">%s</td>
                    </tr>
                    <tr>
                        <th style="border: 1px solid #6A6969;">If interested please click the link given</th>
                        <td style="border: 1px solid #6A6969;"><a href="//127.0.0.1:8000/studentEligibleMark?%s&%s">Click here</a></td>
                    </tr>
                </table>
            </body>
            </html>
            """     % (comp.comp_name, comp.comp_name,comp.comp_profile, comp.ctc, comp.branch, comp.eligibility, comp.date, comp.time, comp.venue, comp.bond, comp.instruction, urllib.parse.urlencode({'token':token}), urllib.parse.urlencode({'stud':user}))
                    email = EmailMessage('Placement', email_body, 'tnpportal7@gmail.com', [user])
                    email.content_subtype = "html"
                    email.send()

            msg = {
                    'success': "Company added and mail sent to eligible students."
                }
            return render(request, 'add_company.html', msg)


@never_cache
def display_company(request):
    if not request.session.get('admin_login', False):
        return HttpResponseRedirect("/login/")
    else:
        types=request.session.get('admin_type')
        if (types=="Super"):
            comp = Company.objects.all().order_by('-id')
            # stud = StudentsEligible.objects.filter(Q(stud_user__in=mailResponse.objects.filter(stud_response=1).values('stud_user')) & Q(comp_name__in=mailResponse.objects.filter(stud_response=1).values('comp_name')))            
            stud = StudentsEligible.objects.raw('SELECT * FROM tnp_admin_studentseligible join tnp_admin_company on tnp_admin_company.comp_name=tnp_admin_studentseligible.comp_name join tnp_admin_mailresponse on tnp_admin_studentseligible.stud_user = tnp_admin_mailresponse.stud_user and tnp_admin_mailresponse.comp_name=tnp_admin_company.comp_name where tnp_admin_mailresponse.stud_response=1')
            # print('expiry age',request.session.get_expiry_age())
            # list_of_student = {}
            # i=0
            # for p in stud:
            #     list_of_student[i]['stud_name'] = p.stud_name
            #     i+=1
            # print(list_of_student)
                
            # students = StudentsEligible.objects.filter(Q(stud_user__in=mailResponse.objects.filter(stud_response=1).values('stud_user')) & Q(comp_name__in=mailResponse.objects.filter(stud_response=1).values('comp_name')))
            if comp.count() > 0:
                if(len(stud)>0):
                    return render(request, 'display_company.html', {'comps': comp, 'eligibles': stud})
                else:
                    return render(request, 'display_company.html', {'comps': comp, 'no_response':True})

            else:
                return render(request, 'display_company.html')
        else:           
            stud = StudentsEligible.objects.raw(f"SELECT * FROM tnp_admin_studentseligible join tnp_admin_company on tnp_admin_company.comp_name=tnp_admin_studentseligible.comp_name join tnp_admin_mailresponse on tnp_admin_studentseligible.stud_user = tnp_admin_mailresponse.stud_user and tnp_admin_mailresponse.comp_name=tnp_admin_company.comp_name where tnp_admin_mailresponse.stud_response=1 and match (tnp_admin_company.branch) against ('{types}') and tnp_admin_studentseligible.branch = '{types}'")
            company = Company.objects.raw(f"SELECT * FROM `tnp_admin_company` where match (tnp_admin_company.branch) against ('{types}')")
            
            if len(company) > 0:
                return render(request, 'display_company.html', {'comps': company, 'eligibles': stud})
            else:
                return render(request, 'display_company.html')
        




@never_cache
def student_placed(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        user = request.GET.get('s')
        company = request.GET.get('c')
        comps = Company.objects.get(comp_name=company)
        student = Resume.objects.get(somaiya_id=user)
        send_mail(
            'Placement',
            'Congratulations you have been placed in ' + company + '.',
            'tnpportal7@gmail.com',
            [user],
            fail_silently=False
        )
        studentPlaced = StudentPlaced(stud_name=student.name, stud_user=user, branch=student.branch, comp_name=company,
                                      ctc=comps.ctc, id_no=student.prn_number)
        studentPlaced.save()
        resumeUpdate = Resume.objects.get(somaiya_id=user)
        if comps.ctc < 600000:
            resumeUpdate.oneto6 = company
            resumeUpdate.save()

            eligible_com = Company.objects.filter(ctc__lte=600000)
            eligible_comp = list(eligible_com.values())
            for e in eligible_comp:
                if StudentsEligible.objects.filter(stud_user=user, comp_name=e['comp_name']).exists():
                    elig = StudentsEligible.objects.get(stud_user=user, comp_name=e['comp_name'])
                    mail = mailResponse.objects.get(stud_user=user, comp_name=e['comp_name'])
                    mail.delete()
                    elig.delete()
        else:
            resumeUpdate.dream = company
            resumeUpdate.save()

            eligible = StudentsEligible.objects.get(stud_user=user, comp_name=company)
            mail = mailResponse.objects.get(stud_user=user, comp_name=company)
            mail.delete()
            eligible.delete()
        return HttpResponseRedirect("/tnp_admin/display_company")


@never_cache
def display_placed(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    elif request.method == "POST":
        types = request.session.get('admin_type')
        if 'filter' in request.POST:
            filter = request.POST['filter']
        else:
            filter = False
        if (types == "Super"):
            if filter == "CTC":
                studentPlaced = StudentPlaced.objects.all().order_by('-ctc')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            elif filter == "Branch":
                studentPlaced = StudentPlaced.objects.all().order_by('branch')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            else:
                studentPlaced = StudentPlaced.objects.all().order_by('comp_name')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
        if (types == "Information Technology"):
            if filter == "CTC":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('-ctc')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            elif filter == "Branch":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('branch')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            else:
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('comp_name')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
        if (types == "Computer"):
            # filter = request.POST['filter']
            if filter == "CTC":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('-ctc')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            elif filter == "Branch":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('branch')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            else:
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('comp_name')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
        if (types == "Electronics & Telecommunication"):
            # filter = request.POST['filter']
            if filter == "CTC":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('-ctc')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            elif filter == "Branch":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('branch')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            else:
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('comp_name')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
        if (types == "Electronics"):
            # filter = request.POST['filter']
            if filter == "CTC":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('-ctc')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            elif filter == "Branch":
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('branch')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})
            else:
                studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('comp_name')
                return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': filter})

    else:
        types = request.session.get('admin_type')
        if (types == "Super"):
            studentPlaced = StudentPlaced.objects.all().order_by('comp_name')
            return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': 'Company'})
        else:
            studentPlaced = StudentPlaced.objects.filter(branch=types).order_by('comp_name')
            return render(request, 'placed_student.html', {'placed': studentPlaced, 'filter': 'Company'})
            
        
#filled students details are all fetched by this and download by these
@never_cache
def filled_student(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if(admin_type =="Super"):

            filled_student_details = Resume.objects.all().order_by('branch')
            
            response = HttpResponse(content_type='text/csv')
            
            response['Content-Disposition'] = 'attachment; filename="student_filled.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Somaiya Id', 'Branch', 'Gender', 'PRN NO.'])
            for filled_student in filled_student_details:
                writer.writerow([filled_student.name,
                filled_student.somaiya_id,filled_student.branch,filled_student.gender,filled_student.prn_number])
        
        elif(admin_type =="Information Technology" ):
            filled_student_details = Resume.objects.filter(branch=admin_type).order_by('branch')
            
            response = HttpResponse(content_type='text/csv')
            
            response['Content-Disposition'] = 'attachment; filename="student_filled.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Somaiya Id', 'Branch', 'Gender', 'PRN NO.'])
            for filled_student in filled_student_details:
                writer.writerow([filled_student.name,
                filled_student.somaiya_id,filled_student.branch,filled_student.gender,filled_student.prn_number])
        
        elif(admin_type=="Computer"):
            filled_student_details = Resume.objects.filter(branch=admin_type).order_by('branch')
            
            response = HttpResponse(content_type='text/csv')
            
            response['Content-Disposition'] = 'attachment; filename="student_filled.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Somaiya Id', 'Branch', 'Gender', 'PRN NO.'])
            for filled_student in filled_student_details:
                writer.writerow([filled_student.name,
                filled_student.somaiya_id,filled_student.branch,filled_student.gender,filled_student.prn_number])
        
        elif(admin_type =="Electronics & Telecommunication" ):
            filled_student_details = Resume.objects.filter(branch=admin_type).order_by('branch')
            
            response = HttpResponse(content_type='text/csv')
            
            response['Content-Disposition'] = 'attachment; filename="student_filled.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Somaiya Id', 'Branch', 'Gender', 'PRN NO.'])
            for filled_student in filled_student_details:
                writer.writerow([filled_student.name,
                filled_student.somaiya_id,filled_student.branch,filled_student.gender,filled_student.prn_number])
        
        elif(admin_type =="Electronics" ):
            filled_student_details = Resume.objects.filter(branch=admin_type).order_by('branch')
            
            response = HttpResponse(content_type='text/csv')
            
            response['Content-Disposition'] = 'attachment; filename="student_filled.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Name', 'Somaiya Id', 'Branch', 'Gender', 'PRN NO.'])
            for filled_student in filled_student_details:
                writer.writerow([filled_student.name,
                filled_student.somaiya_id,filled_student.branch,filled_student.gender,filled_student.prn_number]) 
        return response




@never_cache
def updating_student(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if request.method == 'POST':           
            user_id = request.POST['user_somaiya_id']            
            if('sem 1' in request.POST):
                updating_sem = request.POST['sem 1']
                updating_mark = request.POST['sem 1_marks']
                print('update',updating_sem,updating_mark)
            if('sem 2' in request.POST):
                updating_sem = request.POST['sem 2']
                updating_mark = request.POST['sem 2_marks']
                print('update',updating_sem,updating_mark)
            if('sem 3' in request.POST):
                updating_sem = request.POST['sem 3']
                updating_mark = request.POST['sem 3_marks']
                print('update',updating_sem,updating_mark)
            if('sem 4' in request.POST):
                updating_sem = request.POST['sem 4']
                updating_mark = request.POST['sem 4_marks']
                print('update',updating_sem,updating_mark)
            if('sem 5' in request.POST):
                updating_sem = request.POST['sem 5']
                updating_mark = request.POST['sem 5_marks']
                print('update',updating_sem,updating_mark)
            if('sem 6' in request.POST):
                updating_sem = request.POST['sem 6']
                updating_mark = request.POST['sem 6_marks']
                print('update',updating_sem,updating_mark)
            else:
                print('nothing check')
                return render(request,'update_student.html')
    return render(request,'update_student.html')
        # if(admin_type =="Super"):


@never_cache
def logout_admin(request):
    del request.session['admin_login']
    request.session.modified = True
    return HttpResponseRedirect("/login/")


def pdf(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if(admin_type =="Super"):
            company = request.GET.get('c')
            eligible = StudentsEligible.objects.filter(comp_name=company)
            template = get_template('pdf.html')
            comps = Company.objects.get(comp_name=company)
            # context = {eligible}
            # html = template.render(context)
            pdf = render_to_pdf('eligible_pdf.html', {'context': eligible, 'company': company, 'comps': comps})
        
        else:
            company = request.GET.get('c')
            eligible = StudentsEligible.objects.filter(comp_name=company,branch=admin_type)
            template = get_template('pdf.html')
            comps = Company.objects.get(comp_name=company)
            # context = {eligible}
            # html = template.render(context)
            pdf = render_to_pdf('eligible_pdf.html', {'context': eligible, 'company': company, 'comps': comps})

        return HttpResponse(pdf, content_type='application/pdf')

def company_details(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if (admin_type=="Super"):
            details=Company.objects.all()
            student=StudentPlaced.objects.all()
            it = 0
            extc = 0
            etrx = 0
            comps = 0
            for place in student:
                if place.branch == "Information Technology":
                    it = it + 1
                elif place.branch == "Electronics & Telecommunication":
                    extc = extc + 1
                elif place.branch == "Electronics":
                    etrx = etrx + 1
                else:
                    comps = comps + 1

            pdf=render_to_pdf('company_details.html',{'details':details,'it': it, 'extc': extc, 'etrx': etrx, 'comps': comps,'admin_type':admin_type})
            return HttpResponse(pdf,content_type='application/pdf')
        else:
            details = Company.objects.raw(f"SELECT * FROM `tnp_admin_company` where match (tnp_admin_company.branch) against ('{admin_type}')")
            student=StudentPlaced.objects.filter(branch=admin_type)
            student_placed=0
            for place in student:
                student_placed+=1
            pdf=render_to_pdf('company_details.html',{'details':details,'student':student,'student_placed':student_placed,"admin_type":admin_type})
            return HttpResponse(pdf,content_type='application/pdf')        

def sorting_student(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type = request.session.get('admin_type')
        if (admin_type == "Super"):
            if request.method == "POST":
                student_filter = request.POST['student_filter']

                if student_filter == "Gender":
                    student_details=Resume.objects.all().order_by('gender')
                    return render(request,"sorting_student.html",{"student_details":student_details,'student_filter':student_filter})

                elif student_filter == "Branch":

                    student_details = Resume.objects.all().order_by('branch')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})

                elif student_filter == "Percentage_low":
                    student_details = Resume.objects.all().order_by('agg')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})

                elif student_filter == "Percentage_high":
                    student_details = Resume.objects.all().order_by('-agg')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})

            else:
                return render(request,"sorting_student.html")

        else:
            if request.method == "POST":
                student_filter = request.POST['student_filter']

                if student_filter == "Gender":
                    student_details=Resume.objects.filter(branch=admin_type).order_by('gender')
                    return render(request,"sorting_student.html",{"student_details":student_details,'student_filter':student_filter})

                elif student_filter == "Branch":

                    student_details = Resume.objects.filter(branch=admin_type).order_by('branch')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})

                elif student_filter == "Percentage_low":
                    student_details = Resume.objects.filter(branch=admin_type).order_by('agg')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})

                elif student_filter == "Percentage_high":
                    student_details = Resume.objects.filter(branch=admin_type).order_by('-agg')
                    return render(request,"sorting_student.html",{'student_details':student_details,'student_filter':student_filter})
            else:
                return render(request,"sorting_student.html")
        return render(request,"display_student.html")




def handler404(request, exception):
    return render(request, '404.html', status=404)


def handler500(request):
    return render(request, '500.html', status=500)


def delete_resume(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        types = request.session.get('admin_type')
        if (types=="Super"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(somaiya_id=user)
            resume.delete()
            eligible = StudentsEligible.objects.filter(stud_user=user)
            eligible.delete()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(types=="Information Technology"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(somaiya_id=user,branch=types)
            resume.delete()
            eligible = StudentsEligible.objects.filter(stud_user=user)
            eligible.delete()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(types=="Computer"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(somaiya_id=user,branch=types)
            resume.delete()
            eligible = StudentsEligible.objects.filter(stud_user=user)
            eligible.delete()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(types=="Electronics & Telecommunication"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(somaiya_id=user,branch=types)
            resume.delete()
            eligible = StudentsEligible.objects.filter(stud_user=user)
            eligible.delete()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(types=="Electronics"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(somaiya_id=user,branch=types)
            resume.delete()
            eligible = StudentsEligible.objects.filter(stud_user=user)
            eligible.delete()
            return HttpResponseRedirect("/tnp_admin/user_display")



def unlockResume(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if(admin_type=="Super"):
            user = request.GET.get('s')
            resume = Resume.objects.get(user=user)
            resume.lock = False
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Information Technology"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = False
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Computer"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = False
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Electronics & Telecommunication"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = False
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Electronics"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = False
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")



def lockResume(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')
        if(admin_type=="Super"):
            user = request.GET.get('s')
            resume = Resume.objects.get(user=user)
            resume.lock = True
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Information Technology"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = True
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Computer"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = True
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Electronics & Telecommunication"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = True
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")
        elif(admin_type=="Electronics"):
            user = request.GET.get('s')
            resume = Resume.objects.filter(user=user,branch=admin_type)
            resume.lock = True
            resume.save()
            return HttpResponseRedirect("/tnp_admin/user_display")




def delete_user(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:

        user = request.GET.get('s')
        users = User.objects.filter(username=user)
        users.delete()
        return HttpResponseRedirect("/tnp_admin/user_display")


def delete_company(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        company = request.GET.get('c')
        eligible = StudentsEligible.objects.filter(comp_name=company)
        mail_response = mailResponse.objects.filter(comp_name=company)
        eligible.delete()
        mail_response.delete()
        comp = Company.objects.filter(comp_name=company)
        comp.delete()
        return HttpResponseRedirect("/tnp_admin/display_company")


@never_cache
def add_excel(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        student = request.FILES['excel_student']
        check = student.name
        if check.endswith('.xls') or check.endswith('.xlsx') or check.endswith('.XLS') or check.endswith('.XLSX'):
            # wb = openpyxl.load_workbook(student)
            adding_student = pd.read_excel(student)
            
            msg = {}
            yes,no=0,0
            for student in adding_student.itertuples():
                student_username = student.somaiya_email_id
                stud_name = student.Name
                branch = student.Branch
                # print('studne',pd.isna(student_username)) 
                if User.objects.filter(username=student_username).exists() or pd.isna(stud_name) or pd.isna(student_username) or pd.isna(branch) or (not student_username.endswith('@somaiya.edu')):
                    # print(student)
                    no = no + 1
                else:
                    password = str(''.join(random.choices(string.ascii_uppercase + string.digits, k = 8)))
                    # addUser = User(name=stud_name, username=student_username, password=password, branch=branch)
                    # addUser.save()
                    print('pass',password)
                    send_mail(
                        'Placement Portal',
                        'Id: ' + student_username + '\nPassword: ' + password + '.',
                        'pimsportal@tnpportal.kjsieit.in',
                        [student_username],
                        fail_silently=False,
                    )
                    yes = yes + 1
            msg = {
                "yes": yes,
                "no": no,
                }

            # print('no',no)
                
            # worksheet = wb["Sheet1"]
            # no = 0
            # yes = 0
            # for i,row in enumerate(worksheet.iter_rows()):
            #     row_data = list()
            #     if i == 0:
            #         continue
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data.append(row_data)

            # for add in excel_data:
            #     name = add[0]
            #     username = add[1]
            #     branch = add[2]

            #     if User.objects.filter(username=username).exists() or name == "" or username == "" or branch == "" or (not username.endswith('@somaiya.edu')):
            #         no = no + 1
            #     else:
            #         password = str(''.join(random.choices(string.ascii_uppercase + string.digits, k = 8)))
            #         addUser = User(name=name, username=username, password=password, branch=branch)
            #         addUser.save()
            #         send_mail(
            #             'Placement Portal',
            #             'Id: ' + username + '\nPassword: ' + password + '.',
            #             'tnpportal7@gmail.com',
            #             [username],
            #             fail_silently=False,
            #         )
            #         yes = yes + 1
            # msg = {
            #     "yes": yes,
            #     "no": no,
            #     }
        else:
            msg = {
                "invalidate": "Invalid file format.",
                }
        return render(request, 'add_student.html', msg)

def add_student_from_excel(excel_data):
    yes,no = 0,0
    stud_user = excel_data.somaiya_email_ID
    stud_name = excel_data.student_name
    branch = excel_data.branch
    if(stud_user.endswith('@somaiya.edu')):
        password = str(''.join(random.choices(string.ascii_uppercase + string.digits, k = 8)))
        addUser = User(name=stud_name, username=stud_user, password=password, branch=branch)
        addUser.save()
        # print('pass',password)
        send_mail(
            'Placement Portal',
            'Id: ' + stud_user + '\nPassword: ' + password + '.',
            'pimsportal@tnpportal.kjsieit.in',
            [stud_user],
            fail_silently=False,
        )
        yes += 1
    else:
        no +=1
    return [yes,no]

    # student_username.endswith('@somaiya.edu')
    # print('student',excel_data)


def calculate_percentage(cgpa_score):
    if(cgpa_score<7):
        final_percent =  7.1*cgpa_score + 12
    else:
        final_percent = 7.4*cgpa_score + 12
    return ('%.3f'%final_percent)

@never_cache
def student_data(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        
        if request.method == "POST":
            student_entries_data = request.FILES['student_data']
            checking = student_entries_data.name
            msg = ""
            yes,no=0,0
            if checking.lower().endswith('.xls') or checking.lower().endswith('.xlsx') or checking.lower().endswith('.csv'):
                if(checking.lower().endswith('.xls') or checking.lower().endswith('.xlsx')):
                    student_data = pd.read_excel(student_entries_data)
                else:
                    student_data = pd.read_csv(student_entries_data,encoding='utf-8')
                try:
                    for student in student_data.itertuples():
                        # print('student in data',calculate_percentage(student.average_CGPA))
                        if(User.objects.filter(username=student.somaiya_email_ID).exists()):
                            if(Resume.objects.filter(somaiya_id=student.somaiya_email_ID).exists()):
                                no+=1
                                msg += f"{student.somaiya_email_ID} Data Already exist \n"
                            else:
                                try:
                                    if(not(pd.isna(student.sem_1)) and not(pd.isna(student.sem_2))):
                                        obj = Resume.objects.create(
                                        prn_number=student.prn_no,
                                        name=student.student_name,
                                        college_id=student.college_id_no,
                                        somaiya_id=student.somaiya_email_ID,
                                        non_somiya_id= student.non_somaiya_email_ID,
                                        branch=student.branch,
                                        gender=student.gender,
                                        diploma=None,
                                        sem1=student.sem_1,
                                        sem2=student.sem_2,
                                        ssc_marks=student.marks_10,
                                        hsc_marks=student.marks_12_or_diploma_marks,
                                        sem3=student.sem_3,
                                        sem4=student.sem_4,
                                        sem5=student.sem_5,
                                        sem6= student.sem_6,
                                        mobile_no = student.contact_no,
                                        alternate_phone_no = student.alternate_contact_no,
                                        agg=calculate_percentage(student.average_CGPA),
                                        nearest_railway_station= student.nearest_railway_station,
                                        home_Town=student.home_Town,
                                        average_CGPA=student.average_CGPA,
                                        lock=True
                                        )
                                        # print(obj)
                                        obj.save()
                                    else:
                                        obj = Resume.objects.create(
                                        prn_number=student.prn_no,
                                        college_id=student.college_id_no,
                                        name=student.student_name,
                                        somaiya_id=student.somaiya_email_ID,
                                        non_somiya_id= student.non_somaiya_email_ID,
                                        branch=student.branch,
                                        gender=student.gender,
                                        diploma=student.marks_12_or_diploma_marks,
                                        sem1=None,
                                        sem2=None,
                                        ssc_marks=student.marks_10,
                                        hsc_marks=None,
                                        sem3=student.sem_3,
                                        sem4=student.sem_4,
                                        sem5=student.sem_5,
                                        sem6= student.sem_6,
                                        agg=calculate_percentage(student.average_CGPA),
                                        mobile_no = student.contact_no,
                                        alternate_phone_no = student.alternate_contact_no,
                                        nearest_railway_station= student.nearest_railway_station,
                                        home_Town=student.home_Town,
                                        average_CGPA=student.average_CGPA,
                                        lock=True
                                        )
                                        obj.save()
                                        # print('student is from diploma',student)
                                    # obj = Resume()
                                    yes+=1
                                except Exception as e:
                                    no+=1
                                    msg+="Please fill the information properly"
                                    print('error',e)                            
                                # print('studnet passed from 12',pd.isna(student.marks_12))
                        else:                            
                            add_student_from_excel(student)
                            # if(student_from_excel[0]<0):
                            try:
                                if(not(pd.isna(student.sem_1)) and not(pd.isna(student.sem_2))):
                                    obj = Resume.objects.create(
                                    prn_number=student.prn_no,
                                    name=student.student_name,
                                    college_id=student.college_id_no,
                                    somaiya_id=student.somaiya_email_ID,
                                    non_somiya_id= student.non_somaiya_email_ID,
                                    branch=student.branch,
                                    gender=student.gender,
                                    diploma=None,
                                    sem1=student.sem_1,
                                    sem2=student.sem_2,
                                    ssc_marks=student.marks_10,
                                    hsc_marks=student.marks_12,
                                    sem3=student.sem_3,
                                    sem4=student.sem_4,
                                    sem5=student.sem_5,
                                    sem6= student.sem_6,
                                    mobile_no = student.contact_no,
                                    alternate_phone_no = student.alternate_contact_no,
                                    agg=calculate_percentage(student.average_CGPA),
                                    nearest_railway_station= student.nearest_railway_station,
                                    home_Town=student.home_Town,
                                    average_CGPA=student.average_CGPA,
                                    lock=True
                                    )
                                    # print(obj)
                                    obj.save()
                                else:
                                    obj = Resume.objects.create(
                                    prn_number=student.prn_no,
                                    college_id=student.college_id_no,
                                    name=student.student_name,
                                    somaiya_id=student.somaiya_email_ID,
                                    non_somiya_id= student.non_somaiya_email_ID,
                                    branch=student.branch,
                                    gender=student.gender,
                                    diploma=student.diploma_marks,
                                    sem1=None,
                                    sem2=None,
                                    ssc_marks=student.marks_10,
                                    hsc_marks=None,
                                    sem3=student.sem_3,
                                    sem4=student.sem_4,
                                    sem5=student.sem_5,
                                    sem6= student.sem_6,
                                    agg=calculate_percentage(student.average_CGPA),
                                    mobile_no = student.contact_no,
                                    alternate_phone_no = student.alternate_contact_no,
                                    nearest_railway_station= student.nearest_railway_station,
                                    home_Town=student.home_Town,
                                    average_CGPA=student.average_CGPA,
                                    lock=True
                                    )
                                    obj.save()
                                    # print('student is from diploma',student)
                                # obj = Resume()
                                yes+=1
                            except Exception as e:
                                no+=1
                                msg+="Please fill the information properly"
                                print('error',e)
                            # no+=1
                            # msg += f"Student with email id {student.somaiya_email_ID} does not exist \n"
                except AttributeError:
                    error_occured = 'Something went wrong!'
                    return render(request,'student_data.html',{'error':error_occured})
            else:
                msg += 'File is not a proper format'
            return render(request, 'student_data.html',{'yes_count':yes,"no_count":no,"msg":msg})
        return render(request, 'student_data.html')

@never_cache
def format_of_excel(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        columns = [
        "Sr.no","prn_no",'college_id_no','student_name','branch','gender',
        'somaiya_email_ID','non_somaiya_email_ID','marks_12_or_diploma_marks','marks_10',
        'sem_1','sem_2','sem_3','sem_4','sem_5',"sem_6",
        'average_percentage',"average_CGPA",'Contact no','Alternate contact no',
        'home_Town','nearest_railway_station'
        ]
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename=formatOfStudent.xls'
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet('Sheet1')
        row_nums = 0
        font_style= xlwt.XFStyle()
        font_style.font.bold = True

        for col in range(len(columns)):
            ws.write(row_nums,col,columns[col],font_style)
        
        wb.save(response)

        # df.to_excel("formatOfStudentData.xls")

    return response



@never_cache
def edit_company(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    elif request.GET.get('c'):
        company = request.GET.get('c')
        comp = Company.objects.filter(comp_name=company)
        if comp.exists():
            data = {
                'comp': comp,
                }
            return render(request, 'edit_company.html', data)
        else:
            return HttpResponseRedirect("/tnp_admin/display_company")
    else:
        c_name = request.POST['c_name'].strip()
        compCheck = Company.objects.filter(comp_name=c_name)

        if compCheck.exists():
            c_profile = request.POST['c_profile']
            eligible=request.POST['eligible']
            bond = request.POST['bond']
            date = request.POST['date']
            time = request.POST['time']
            venue = request.POST['venue']
            branch = request.POST['branch']
            instruction = request.POST['instruction']
            campus = request.POST['campus']

            compUpdate = Company.objects.get(comp_name=c_name)
            compUpdate.comp_profile = c_profile
            compUpdate.eligibility=eligible
            compUpdate.bond = bond
            compUpdate.date = date
            compUpdate.time = time
            compUpdate.venue = venue
            compUpdate.instruction = instruction
            compUpdate.campus = campus

            compUpdate.save()
            return HttpResponseRedirect("/tnp_admin/display_company")
        else:
            return HttpResponseRedirect("/tnp_admin/display_company")


@never_cache
def endTerm(request):
    if not request.session.get('admin_login'):
        return HttpResponseRedirect("/login/")
    else:
        admin_type=request.session.get('admin_type')

        studentPlaced = StudentPlaced.objects.all().order_by('branch').order_by('ctc').values_list('stud_name', 'branch', 'id_no', 'stud_user', 'comp_name', 'ctc')

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="placements.csv"'
        writer = csv.writer(response)
        writer.writerow(['Name', 'Branch', 'Id', 'Username', 'Company', 'CTC'])
        for placed in studentPlaced:
            writer.writerow(placed)
        User.objects.all().delete()
        Resume.objects.all().delete()
        Company.objects.all().delete()
        StudentsEligible.objects.all().delete()
        StudentPlaced.objects.all().delete()
        return response